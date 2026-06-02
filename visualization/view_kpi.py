"""
view_kpi.py — Rapport KPI multi-fichiers

Scanne le dossier output pour tous les *_results_*.csv.
Calcule les KPIs globaux + détail par fichier en accordéon cliquable.
Génère kpi_report_<ts>.xlsx dans le dossier de sortie.

Usage standalone :
    python visualization/view_kpi.py <output_dir>
"""

import sys
import os

_DIR  = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_DIR)
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

import base64
import json
from pathlib import Path

import numpy as np
import pandas as pd
import dash
from dash import html

PAGE_ZOOM = 0.9
FLOAT_TOL = 1e-6

_KPI_CACHE_FILE = "_kpi_cache.json"


class _NpEncoder(json.JSONEncoder):
    """Sérialise les types numpy (int64, float64) en types Python natifs."""
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)


def _save_kpi_cache(output_dir: str, rows_by_file: dict) -> None:
    """Sauvegarde rows_by_file dans _kpi_cache.json pour les affichages futurs."""
    try:
        path = os.path.join(output_dir, _KPI_CACHE_FILE)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(rows_by_file, f, cls=_NpEncoder, ensure_ascii=False,
                      separators=(",", ":"))
    except Exception:
        pass


def _load_kpi_cache(output_dir: str, current_file_names: set) -> dict | None:
    """Charge le cache JSON si les fichiers du dossier correspondent exactement."""
    try:
        path = os.path.join(output_dir, _KPI_CACHE_FILE)
        if not os.path.exists(path):
            return None
        with open(path, "r", encoding="utf-8") as f:
            cached = json.load(f)
        if set(cached.keys()) == current_file_names:
            return cached
        return None
    except Exception:
        return None


# ── Logo helper ────────────────────────────────────────────────────────────────

def _load_logo(filename: str) -> str:
    path = os.path.join(_ROOT, filename)
    if not os.path.exists(path):
        return ""
    ext  = filename.rsplit(".", 1)[-1].lower()
    mime = "jpeg" if ext in ("jpg", "jpeg") else ext
    with open(path, "rb") as f:
        return f"data:image/{mime};base64," + base64.b64encode(f.read()).decode()


# ── Metric helpers ─────────────────────────────────────────────────────────────

def _cog_offset(df_p: pd.DataFrame) -> float:
    tw = df_p["weight"].sum()
    if tw <= 0:
        return 0.0
    cx = (df_p["weight"] * (df_p["x"] + df_p["length"] / 2)).sum() / tw
    cy = (df_p["weight"] * (df_p["y"] + df_p["width"]  / 2)).sum() / tw
    dx = cx - float(df_p["pallet_length"].iloc[0]) / 2
    dy = cy - float(df_p["pallet_width"].iloc[0])  / 2
    return (dx**2 + dy**2) ** 0.5


def _surf_fill(df_p: pd.DataFrame) -> float:
    p_len = int(round(float(df_p["pallet_length"].iloc[0])))
    p_wid = int(round(float(df_p["pallet_width"].iloc[0])))
    p_area = p_len * p_wid
    if p_area <= 0:
        return 0.0
    xs = df_p["x"].values; ys = df_p["y"].values
    ls = df_p["length"].values; ws = df_p["width"].values
    x0a = np.clip(xs.astype(int), 0, p_len)
    x1a = np.clip(np.round(xs + ls).astype(int), 0, p_len)
    y0a = np.clip(ys.astype(int), 0, p_wid)
    y1a = np.clip(np.round(ys + ws).astype(int), 0, p_wid)
    grid = np.zeros((p_len, p_wid), dtype=bool)
    for i in range(len(xs)):
        if x1a[i] > x0a[i] and y1a[i] > y0a[i]:
            grid[x0a[i]:x1a[i], y0a[i]:y1a[i]] = True
    return float(grid.sum()) / p_area


def _per_pallet_rows(df: pd.DataFrame) -> list:
    rows = []
    for pid in sorted(df["pallet_id"].unique()):
        df_p    = df[df["pallet_id"] == pid]
        clients = sorted(df_p["client_id"].unique())
        tw      = df_p["weight"].sum()
        fill    = float(df_p["volumetric_fill_ratio"].iloc[0]) \
                  if "volumetric_fill_ratio" in df_p.columns else 0.0
        sf      = _surf_fill(df_p)
        cog     = _cog_offset(df_p)
        cog_x   = float((df_p["weight"] * (df_p["x"] + df_p["length"] / 2)).sum() / tw) if tw > 0 else 0.0
        cog_y   = float((df_p["weight"] * (df_p["y"] + df_p["width"]  / 2)).sum() / tw) if tw > 0 else 0.0
        cog_z   = float((df_p["weight"] * (df_p["z"] + df_p["height"] / 2)).sum() / tw) if tw > 0 else 0.0
        stab    = float(df_p["worst_stability_ratio"].iloc[0]) \
                  if "worst_stability_ratio" in df_p.columns else 0.0
        rows.append({
            "pid": pid, "multi": len(clients) > 1, "clients": clients,
            "fill": fill, "surf_fill": sf,
            "cog": cog, "cog_x": cog_x, "cog_y": cog_y, "cog_z": cog_z,
            "height": float((df_p["z"] + df_p["height"]).max()),
            "p1": int((df_p["priority"] == 1).sum()),
            "p2": int((df_p["priority"] == 2).sum()),
            "weight": float(tw), "n_boxes": len(df_p), "stability": stab,
        })
    return rows


def _load_all_results(output_dir: str) -> dict:
    p = Path(output_dir)
    files = sorted(p.glob("*_results_*.csv"), key=lambda f: f.stat().st_mtime, reverse=True)
    result = {}
    for f in files:
        try:
            df = pd.read_csv(f, sep=";")
            for col in ["x", "y", "z", "length", "width", "height",
                        "weight", "pallet_length", "pallet_width", "pallet_height"]:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce")
            df["pallet_id"] = df["pallet_id"].astype(int)
            df["client_id"] = df["client_id"].astype(int)
            df["priority"]  = df["priority"].astype(int)
            result[f.name] = df
        except Exception:
            pass
    return result


# ── Layout builders ────────────────────────────────────────────────────────────

def _kpi_chip(label: str, value: str, color: str) -> html.Div:
    return html.Div([
        html.Div(label, style={"fontSize": "13px", "color": "#9ca3af", "lineHeight": "1.2"}),
        html.Div(value, style={"fontSize": "17px", "fontWeight": "700",
                                "color": color, "lineHeight": "1.3"}),
    ], style={"textAlign": "center", "minWidth": "94px", "padding": "0 5px", "flexShrink": "0"})


def _build_global_kpis(rows_by_file: dict) -> html.Div:
    if not rows_by_file:
        return html.Div("Aucun fichier de résultats trouvé dans le dossier de sortie.",
                        style={"color": "#9ca3af", "padding": "20px", "textAlign": "center"})

    total_palettes = total_multi = total_boxes = total_p1 = total_p2 = 0
    fill_vals: list = []; surf_vals: list = []

    for rows in rows_by_file.values():
        total_palettes += len(rows)
        total_multi    += sum(1 for r in rows if r["multi"])
        total_boxes    += sum(r["n_boxes"] for r in rows)
        total_p1       += sum(r["p1"] for r in rows)
        total_p2       += sum(r["p2"] for r in rows)
        fill_vals.extend(r["fill"]      for r in rows)
        surf_vals.extend(r["surf_fill"] for r in rows)

    avg_fill   = sum(fill_vals) / len(fill_vals) if fill_vals else 0.0
    avg_surf   = sum(surf_vals) / len(surf_vals) if surf_vals else 0.0
    ratio      = total_p2 / total_p1 if total_p1 > 0 else float("inf")
    multi_rate = total_multi / total_palettes if total_palettes else 0.0
    boxes_per_pal = total_boxes / total_palettes if total_palettes else 0.0
    p1_per_pal    = total_p1    / total_palettes if total_palettes else 0.0
    p2_per_pal    = total_p2    / total_palettes if total_palettes else 0.0

    kpi_style = {"backgroundColor": "white", "borderRadius": "8px",
                 "padding": "14px 18px", "boxShadow": "0 1px 4px rgba(0,0,0,0.08)",
                 "textAlign": "center", "minWidth": "110px"}
    lbl = {"fontSize": "12px", "color": "#9ca3af", "marginBottom": "4px"}
    val = lambda c: {"fontSize": "30px", "fontWeight": "bold", "color": c, "lineHeight": "1.1"}
    sub = {"fontSize": "12px", "color": "#6b7280", "marginTop": "4px"}

    kpis = [
        ("Fichiers analysés",      str(len(rows_by_file)), "#1d4ed8", None),
        ("Total palettes",         str(total_palettes),   "#374151", None),
        ("Multi-client",           str(total_multi),      "#dc2626" if total_multi > 0 else "#374151", None),
        ("Taux multi",             f"{multi_rate:.1%}",   "#dc2626" if multi_rate > 0 else "#374151", None),
        ("Rempli. vol. moy.",      f"{avg_fill:.1%}",     "#16a34a", None),
        ("Rempli. surf. moy.",     f"{avg_surf:.1%}",     "#0d9488", None),
        ("Total Articles",         str(total_boxes),      "#374151", f"{boxes_per_pal:.2f} / palette"),
        ("Total P1 (Meubles)",     str(total_p1),         "#374151", f"{p1_per_pal:.2f} / palette"),
        ("Total P2 (Colis)",       str(total_p2),         "#ea580c", f"{p2_per_pal:.2f} / palette"),
        ("Ratio P2 / P1",          f"{ratio:.2f}" if ratio != float('inf') else "∞",
                                                           "#dc2626" if ratio > 0.5 else "#16a34a", None),
    ]
    return html.Div(
        style={"display": "flex", "gap": "12px", "flexWrap": "wrap",
               "justifyContent": "center", "marginBottom": "28px"},
        children=[
            html.Div(style=kpi_style, children=[
                html.Div(label, style=lbl),
                html.Div(value, style=val(color)),
                html.Div(subtext, style=sub) if subtext else None,
            ])
            for label, value, color, subtext in kpis
        ],
    )


def _build_file_section(filename: str, rows: list, csv_path: str = None) -> html.Details:
    if not rows:
        return html.Details([html.Summary(filename)], style={"marginBottom": "8px"})

    has_links = bool(csv_path)
    n_total = len(rows); n_multi = sum(1 for r in rows if r["multi"])
    avg_fill = sum(r["fill"] for r in rows) / n_total
    avg_surf = sum(r["surf_fill"] for r in rows) / n_total
    t_boxes = sum(r["n_boxes"] for r in rows)
    t_p1 = sum(r["p1"] for r in rows); t_p2 = sum(r["p2"] for r in rows)
    ratio = t_p2 / t_p1 if t_p1 > 0 else float("inf")

    mono_fills = [r["fill"] for r in rows if not r["multi"]]
    if mono_fills:
        mean_mono = sum(mono_fills) / len(mono_fills)
        std_mono  = float(np.std(mono_fills))
    else:
        mean_mono = None; std_mono = None

    def _fill_color(v):
        return "#16a34a" if v >= 0.60 else "#f59e0b" if v >= 0.40 else "#dc2626"

    multi_rate = n_multi / n_total if n_total else 0.0

    chips = html.Div([
        _kpi_chip("Palettes",      str(n_total),                "#1d4ed8"),
        _kpi_chip("Multi-cl.",     str(n_multi),                "#dc2626" if n_multi > 0 else "#9ca3af"),
        _kpi_chip("Taux multi",    f"{multi_rate:.1%}",         "#dc2626" if multi_rate > 0 else "#9ca3af"),
        _kpi_chip("Rempli. vol.",  f"{avg_fill:.1%}",           "#16a34a" if avg_fill >= 0.6 else "#f59e0b"),
        _kpi_chip("Rempli. surf.", f"{avg_surf:.1%}",           "#0d9488"),
        _kpi_chip("Moy. rempli. Mono",
                  f"{mean_mono:.1%}" if mean_mono is not None else "—",
                  _fill_color(mean_mono) if mean_mono is not None else "#9ca3af"),
        _kpi_chip("Écart-type Mono", f"{std_mono:.1%}" if std_mono is not None else "—", "#374151"),
        _kpi_chip("Articles",      str(t_boxes),                "#374151"),
        _kpi_chip("P1 (Meubles)",  str(t_p1),                  "#374151"),
        _kpi_chip("P2 (Colis)",    str(t_p2),                  "#ea580c"),
        _kpi_chip("Ratio P2/P1",   f"{ratio:.2f}" if ratio != float('inf') else "∞",
                  "#dc2626" if ratio > 0.5 else "#16a34a"),
    ], style={"display": "flex", "gap": "8px", "alignItems": "center",
              "flexShrink": "0", "flexWrap": "nowrap"})

    if has_links:
        from urllib.parse import quote as _quote
        _grid_href = f"open-grid?csv={_quote(csv_path, safe='')}"
        open_btn = html.A(
            "🖥 Ouvrir ↗",
            href=_grid_href,
            target="_blank",
            style={"background": "#2563eb", "color": "white", "borderRadius": "5px",
                   "padding": "3px 10px", "fontSize": "12px", "fontWeight": "600",
                   "cursor": "pointer", "flexShrink": "0", "textDecoration": "none",
                   "display": "inline-block"},
        )
    else:
        open_btn = html.Span()

    summary_content = html.Div([
        html.Span(filename, style={"fontWeight": "600", "fontSize": "15px", "color": "#1e293b",
                                   "flex": "1", "minWidth": "0", "overflow": "hidden",
                                   "textOverflow": "ellipsis", "whiteSpace": "nowrap"}),
        open_btn,
        chips,
    ], style={"display": "inline-flex", "alignItems": "center", "gap": "12px",
              "width": "calc(100% - 28px)", "verticalAlign": "middle", "flexWrap": "nowrap"})

    th = {"padding": "8px 10px", "textAlign": "left", "borderBottom": "2px solid #e2e8f0",
          "color": "#6b7280", "fontSize": "12px", "whiteSpace": "nowrap",
          "position": "sticky", "top": "0", "backgroundColor": "#f8fafc"}
    td = {"padding": "6px 10px", "fontSize": "13px", "borderBottom": "1px solid #f1f5f9",
          "whiteSpace": "nowrap"}

    # Colonne vide en tête si liens actifs (bouton zoom collé au nom de palette)
    headers = ([""] if has_links else []) + [
        "Palette", "Client(s)", "Rempli. vol.", "Rempli. surf.", "Poids (kg)",
        "Colis", "P1", "P2", "Hauteur (cm)", "CdG X (cm)", "CdG Y (cm)",
        "CdG Z (cm)", "H / Rempli.", "Ratio stabilité",
    ]

    _td_btn = {"padding": "4px 6px", "borderBottom": "1px solid #f1f5f9", "whiteSpace": "nowrap"}

    table_rows = []
    for i, r in enumerate(rows):
        bg = "#fff8f8" if r["multi"] else ("white" if i % 2 == 0 else "#f9fafb")

        if has_links:
            from urllib.parse import quote as _quote
            _zoom_href = f"open-zoom?csv={_quote(csv_path, safe='')}&pid={r['pid']}"
            zoom_cell = html.Td(
                html.A("🔍", href=_zoom_href, target="_blank",
                       title=f"Ouvrir Palette {r['pid']} en Vue Zoom",
                       style={"background": "#0891b2", "color": "white",
                              "borderRadius": "4px", "padding": "2px 7px",
                              "fontSize": "12px", "cursor": "pointer",
                              "textDecoration": "none", "display": "inline-block"}),
                style=_td_btn,
            )
            first_cells = [zoom_cell]
        else:
            first_cells = []

        cells = first_cells + [
            html.Td(f"Palette {r['pid']}", style={**td, "fontWeight": "600"}),
            html.Td("Multi" if r["multi"] else str(r["clients"][0]), style=td),
            html.Td(f"{r['fill']:.1%}", style={**td,
                "color": "#16a34a" if r["fill"] >= 0.6 else "#f59e0b" if r["fill"] >= 0.4 else "#dc2626"}),
            html.Td(f"{r['surf_fill']:.1%}", style={**td,
                "color": "#0d9488" if r["surf_fill"] >= 0.6 else "#f59e0b" if r["surf_fill"] >= 0.4 else "#dc2626"}),
            html.Td(f"{r['weight']:.1f}", style=td),
            html.Td(str(r["n_boxes"]), style=td),
            html.Td(str(r["p1"]), style=td),
            html.Td(str(r["p2"]), style={**td, "color": "#ea580c" if r["p2"] > 0 else "#9ca3af"}),
            html.Td(f"{r['height']:.1f}", style=td),
            html.Td(f"{r['cog_x']:.1f}", style=td), html.Td(f"{r['cog_y']:.1f}", style=td),
            html.Td(f"{r['cog_z']:.1f}", style=td),
            html.Td(f"{(r['height'] / r['fill']):.0f}" if r['fill'] > 0 else "—", style=td),
            html.Td(f"{r['stability']:.2f}", style={**td,
                "color": "#dc2626" if r["stability"] > 5.0 else "#f59e0b" if r["stability"] > 3.0 else "#16a34a"}),
        ]
        table_rows.append(html.Tr(style={"backgroundColor": bg}, children=cells))

    table = html.Table(style={"width": "100%", "borderCollapse": "collapse"},
                       children=[html.Thead(html.Tr([html.Th(h, style=th) for h in headers])),
                                  html.Tbody(table_rows)])

    return html.Details(
        style={"border": "1px solid #e2e8f0", "borderRadius": "8px",
               "marginBottom": "10px", "overflow": "hidden", "backgroundColor": "white"},
        children=[
            html.Summary(summary_content,
                         style={"padding": "11px 14px", "cursor": "pointer",
                                "background": "#f8fafc", "userSelect": "none", "fontSize": "19px"}),
            html.Div(style={"overflowX": "auto", "maxHeight": "400px", "overflowY": "auto"},
                     children=[table]),
        ],
    )


def build_kpi_layout(output_dir: str, logo_b64: str, logo2_b64: str,
                     rows_by_file: dict = None) -> html.Div:
    if rows_by_file is None:
        # Données non fournies : charger depuis cache ou recalculer
        current_names = {f.name for f in Path(output_dir).glob("*_results_*.csv")}
        rows_by_file  = _load_kpi_cache(output_dir, current_names)
        if rows_by_file is None:
            all_data     = _load_all_results(output_dir)
            rows_by_file = {fname: _per_pallet_rows(df) for fname, df in all_data.items()}
            _save_kpi_cache(output_dir, rows_by_file)
    # all_data sert uniquement pour len() et l'itération des noms
    all_data = {fname: None for fname in rows_by_file}

    header_children = [
        html.Img(src=logo_b64, style={"height": "68px", "objectFit": "contain"}) if logo_b64 else html.Div(),
        html.Div(style={"position": "absolute", "width": "100%", "textAlign": "center",
                        "pointerEvents": "none"},
                 children=[html.H2("Rapport KPI", style={"color": "#333", "margin": "0"})]),
    ]
    if logo2_b64:
        header_children.append(html.Img(src=logo2_b64,
                                         style={"height": "68px", "objectFit": "contain", "marginLeft": "auto"}))
    header = html.Div(style={"display": "flex", "alignItems": "center",
                             "marginBottom": "8px", "position": "relative"},
                      children=header_children)

    n_files = len(all_data)
    file_sections = (
        [_build_file_section(fname, rows_by_file[fname],
                             csv_path=os.path.join(output_dir, fname))
         for fname in all_data]
        if all_data else
        [html.Div("Aucun fichier *_results_*.csv trouvé dans le dossier de sortie.",
                  style={"color": "#9ca3af", "padding": "20px", "textAlign": "center",
                         "background": "white", "borderRadius": "8px"})]
    )
    detail_title = html.H3(f"Détail par fichier ({n_files} résultat{'s' if n_files > 1 else ''})",
                           style={"color": "#333", "marginBottom": "12px"})

    return html.Div(
        style={"fontFamily": "Arial, sans-serif", "backgroundColor": "#f5f5f5",
               "minHeight": "100vh", "padding": "16px"},
        children=[
            header,
            html.P(f"Dossier analysé : {output_dir}",
                   style={"textAlign": "center", "color": "#9ca3af",
                          "marginTop": "0", "marginBottom": "12px", "fontSize": "13px"}),
            html.Hr(style={"borderColor": "#ddd", "marginBottom": "20px"}),
            _build_global_kpis(rows_by_file),
            detail_title,
            html.Div(file_sections),
        ],
    )


# ── Dash app builder ───────────────────────────────────────────────────────────

def build_kpi_app(output_dir: str) -> dash.Dash:
    logo_b64  = _load_logo("logo_fournier.png")
    logo2_b64 = _load_logo("logo_u4log.jpg")

    _prefix = os.environ.get("PALLET_KPI_PREFIX", "")
    if _prefix and _prefix != "/":
        app = dash.Dash(__name__, title="Rapport KPI",
                        routes_pathname_prefix=_prefix,
                        requests_pathname_prefix=_prefix)
    else:
        app = dash.Dash(__name__, title="Rapport KPI")
    app.layout = build_kpi_layout(output_dir, logo_b64, logo2_b64)

    app.index_string = app.index_string.replace(
        "<head>",
        f"<head><style>body {{ zoom: {PAGE_ZOOM}; }}</style>"
        f"<script>window.addEventListener('pageshow', function(e) {{"
        f"  if (e.persisted) {{ window.location.reload(true); }}"
        f"}});</script>",
    )

    @app.server.after_request
    def _no_cache(response):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"]        = "no-cache"
        response.headers["Expires"]       = "0"
        return response

    return app


# ── Excel export ───────────────────────────────────────────────────────────────

def generate_excel_report(output_dir: str, excel_path: str = None) -> str:
    """Génère kpi_report_<ts>.xlsx dans output_dir. Retourne le chemin créé."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[KPI] openpyxl non installé — Excel non généré.\n"
              "      Installez avec : pip install openpyxl")
        return ""

    all_data = _load_all_results(output_dir)
    if not all_data:
        print("[KPI] Aucun fichier de résultats — Excel non généré.")
        return ""

    if excel_path is None:
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = str(Path(output_dir) / f"kpi_report_{ts}.xlsx")

    FMT_PCT = "0.0%"; FMT_FLOAT = "0.000"; FMT_INT = "0"

    wb         = openpyxl.Workbook()
    hdr_fill   = PatternFill("solid", fgColor="1E293B")
    sub_fill   = PatternFill("solid", fgColor="334155")
    title_fill = PatternFill("solid", fgColor="0F172A")
    hdr_font   = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=14)

    ws       = wb.active
    ws.title = "Rapport KPI"

    def _write_row(values, formats=None, *, header=False):
        ws.append(values)
        row_num = ws.max_row
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            if header:
                cell.fill = hdr_fill; cell.font = hdr_font
                cell.alignment = Alignment(wrap_text=True, horizontal="center")
            elif formats and col_idx - 1 < len(formats) and formats[col_idx - 1]:
                cell.number_format = formats[col_idx - 1]

    rows_by_file = {fname: _per_pallet_rows(df) for fname, df in all_data.items()}
    _save_kpi_cache(output_dir, rows_by_file)   # cache pour affichage Dash rapide

    total_palettes = total_multi = total_boxes = total_p1 = total_p2 = 0
    fill_vals: list = []; surf_vals: list = []
    for rows in rows_by_file.values():
        total_palettes += len(rows); total_multi += sum(1 for r in rows if r["multi"])
        total_boxes += sum(r["n_boxes"] for r in rows)
        total_p1 += sum(r["p1"] for r in rows); total_p2 += sum(r["p2"] for r in rows)
        fill_vals.extend(r["fill"] for r in rows); surf_vals.extend(r["surf_fill"] for r in rows)

    avg_fill   = sum(fill_vals) / len(fill_vals) if fill_vals else 0.0
    avg_surf   = sum(surf_vals) / len(surf_vals) if surf_vals else 0.0
    ratio_glob = total_p2 / total_p1 if total_p1 > 0 else 0.0
    multi_rate = total_multi / total_palettes if total_palettes else 0.0

    title_cell = ws.cell(row=1, column=1, value="Rapport KPI")
    title_cell.fill = title_fill; title_cell.font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    ws.append([])

    glob_section = ws.cell(row=ws.max_row + 1, column=1, value="KPIs Globaux")
    glob_section.fill = sub_fill; glob_section.font = hdr_font
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=13)

    boxes_per_pal = total_boxes / total_palettes if total_palettes else 0.0
    p1_per_pal    = total_p1    / total_palettes if total_palettes else 0.0
    p2_per_pal    = total_p2    / total_palettes if total_palettes else 0.0

    _write_row(["Fichiers analysés", "Total palettes", "Multi-client",
                "Taux multi", "Rempli. vol. moy.", "Rempli. surf. moy.",
                "Total Articles", "Articles / palette",
                "Total P1 (Meubles)", "P1 / palette",
                "Total P2 (Colis)", "P2 / palette", "Ratio P2/P1"], header=True)
    _write_row([len(all_data), total_palettes, total_multi,
                multi_rate, avg_fill, avg_surf,
                total_boxes, boxes_per_pal, total_p1, p1_per_pal, total_p2, p2_per_pal, ratio_glob],
               formats=[FMT_INT]*3 + [FMT_PCT]*3 + [FMT_INT, FMT_FLOAT]*2 + [FMT_INT, FMT_FLOAT, FMT_FLOAT])
    ws.append([]); ws.append([])

    detail_section = ws.cell(row=ws.max_row + 1, column=1, value="Détail par fichier")
    detail_section.fill = sub_fill; detail_section.font = hdr_font
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=13)

    _write_row(["Fichier", "Palettes", "Multi-client", "Taux multi",
                "Rempli. vol.", "Rempli. surf.", "Moy. rempli. Mono", "Écart-type Mono",
                "Articles", "P1 (Meubles)", "P2 (Colis)", "Ratio P2/P1"], header=True)

    for fname, rows in rows_by_file.items():
        if not rows: continue
        n_total = len(rows); n_multi = sum(1 for r in rows if r["multi"])
        t_boxes = sum(r["n_boxes"] for r in rows)
        t_p1 = sum(r["p1"] for r in rows); t_p2 = sum(r["p2"] for r in rows)
        af = sum(r["fill"] for r in rows) / n_total
        asf = sum(r["surf_fill"] for r in rows) / n_total
        rr = t_p2 / t_p1 if t_p1 > 0 else 0.0
        mr = n_multi / n_total if n_total else 0.0
        mono_fills = [r["fill"] for r in rows if not r["multi"]]
        mean_mono = sum(mono_fills) / len(mono_fills) if mono_fills else None
        std_mono  = float(np.std(mono_fills)) if mono_fills else None
        _write_row([fname, n_total, n_multi, mr, af, asf,
                    mean_mono if mean_mono is not None else "",
                    std_mono  if std_mono  is not None else "",
                    t_boxes, t_p1, t_p2, rr],
                   formats=[None, FMT_INT, FMT_INT, FMT_PCT, FMT_PCT, FMT_PCT,
                             FMT_PCT, FMT_PCT, FMT_INT, FMT_INT, FMT_INT, FMT_FLOAT])

    ws.column_dimensions["A"].width = 42
    for idx in range(2, 14):
        ws.column_dimensions[get_column_letter(idx)].width = 16

    ws2 = wb.create_sheet(title="Détail par palette")
    pal_headers = ["Nom fichier", "Palette", "Client(s)",
                   "Rempli. vol.", "Rempli. surf.", "Poids (kg)",
                   "Colis", "P1", "P2", "Hauteur (cm)",
                   "CdG X (cm)", "CdG Y (cm)", "CdG Z (cm)", "H / Rempli.", "Ratio stabilité"]
    pal_formats = [None, None, None, FMT_PCT, FMT_PCT, FMT_FLOAT,
                   FMT_INT, FMT_INT, FMT_INT,
                   FMT_FLOAT, FMT_FLOAT, FMT_FLOAT, FMT_FLOAT, FMT_FLOAT, FMT_FLOAT]
    ws2.append(pal_headers)
    for col_idx in range(1, len(pal_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    def _fill_color(v): return "16A34A" if v >= 0.6 else "F59E0B" if v >= 0.4 else "DC2626"
    def _surf_color(v): return "0D9488" if v >= 0.6 else "F59E0B" if v >= 0.4 else "DC2626"
    def _stab_color(v): return "DC2626" if v > 5.0 else "F59E0B" if v > 3.0 else "16A34A"

    for fname, rows in rows_by_file.items():
        if not rows: continue
        first_row_for_file = ws2.max_row + 1
        for r in rows:
            client_label = "Multi" if r["multi"] else str(r["clients"][0])
            h_over_fill  = (r["height"] / r["fill"]) if r["fill"] > 0 else None
            values = [fname, f"Palette {r['pid']}", client_label,
                      r["fill"], r["surf_fill"], round(r["weight"], 1),
                      r["n_boxes"], r["p1"], r["p2"], round(r["height"], 1),
                      round(r["cog_x"], 1), round(r["cog_y"], 1), round(r["cog_z"], 1),
                      round(h_over_fill, 0) if h_over_fill is not None else "",
                      round(r["stability"], 3)]
            ws2.append(values)
            row_num = ws2.max_row
            for col_idx, fmt in enumerate(pal_formats, start=1):
                if fmt: ws2.cell(row=row_num, column=col_idx).number_format = fmt
            ws2.cell(row=row_num, column=4).font = Font(color=_fill_color(r["fill"]))
            ws2.cell(row=row_num, column=5).font = Font(color=_surf_color(r["surf_fill"]))
            if r["p2"] > 0: ws2.cell(row=row_num, column=9).font = Font(color="EA580C")
            ws2.cell(row=row_num, column=15).font = Font(color=_stab_color(r["stability"]))
        last_row_for_file = ws2.max_row
        if last_row_for_file > first_row_for_file:
            ws2.merge_cells(start_row=first_row_for_file, start_column=1,
                            end_row=last_row_for_file, end_column=1)
            ws2.cell(row=first_row_for_file, column=1).alignment = Alignment(vertical="center")

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    for idx in range(4, len(pal_headers) + 1):
        ws2.column_dimensions[get_column_letter(idx)].width = 14
    ws2.freeze_panes = "B2"

    wb.save(excel_path)
    print(f"[KPI] Rapport Excel généré : {excel_path}")
    return excel_path


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage : python visualization/view_kpi.py <output_dir>")
        sys.exit(1)

    output_dir = sys.argv[1]
    if not os.path.isdir(output_dir):
        print(f"[KPI] Dossier introuvable : {output_dir}")
        sys.exit(1)

    app  = build_kpi_app(output_dir)
    host = os.environ.get("PALLET_HOST",     "127.0.0.1")
    port = int(os.environ.get("PALLET_KPI_PORT", "8052"))
    base = f"http://{host}:{port}"

    print(f"\n[KPI] Serveur disponible sur : {base}")
    print("[KPI] Ctrl+C pour arrêter.\n")

    app.run(debug=False, host=host, port=port)


if __name__ == "__main__":
    main()


# ── Integrated mode (used by visualizer.py) ───────────────────────────────────

def kpi_layout(output_dir: str, rows_by_file: dict = None) -> html.Div:
    """Retourne le layout KPI. Si rows_by_file est fourni, saute le rechargement CSV."""
    logo_b64  = _load_logo("logo_fournier.png")
    logo2_b64 = _load_logo("logo_u4log.jpg")
    return build_kpi_layout(output_dir, logo_b64, logo2_b64, rows_by_file=rows_by_file)


def register(app: dash.Dash, state: dict) -> None:
    """Les liens KPI → Vue utilisent des routes Flask directes (open-grid / open-zoom).
    Aucun callback Dash n'est nécessaire ici."""
    pass
