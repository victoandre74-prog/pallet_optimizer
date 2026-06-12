"""
run_sweep_2.py — Optimisation Pareto (NSGA-II via Optuna) des paramètres multi-client.

Objectifs minimisés simultanément :
    1. final_pallets      — nombre total de palettes non vides
    2. multi_client_ratio — n_multi_final / final_pallets  (taux de palettes multi-AR)

Paramètres explorés :
    multi_client_minimum_ratio  ∈ [0.02, 0.23]   (soft-stop lower bound)
    mc_max_delta                ∈ [0.01, 0.25]   → mc_max = mc_min + delta  (garantit mc_max > mc_min)
    min_filling_ratio           ∈ [0.20, 0.55]   (seuil remplissage régime ≤10 palettes)

Usage :
    python run_sweep_2.py [--trials 80] [--workers 4] [--seed 42]

Output :
    sweep2_results.csv    tous les essais, flag pareto
    sweep2_results.xlsx   rapport avec front de Pareto mis en évidence
"""

import argparse
import copy
import glob
import io
import os
import re
import sys
import time
import csv
from concurrent.futures import ProcessPoolExecutor, as_completed

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Path setup ─────────────────────────────────────────────────────────────────
_DIR  = os.path.dirname(os.path.abspath(__file__))            # .../pallet_optimizer/dev/experiments/
_ROOT = os.path.dirname(os.path.dirname(_DIR))                 # .../pallet_optimizer/ (git root)
_SRC  = os.path.join(_ROOT, "src")
if _SRC not in sys.path:
    sys.path.insert(0, _SRC)

from config.parameters import OptimizationParameters
from file_io.csv_reader import read_boxes_from_csv
from optimizer.pallet_optimizer import optimize_palletization
from heuristics.post_processing import postprocess

from datetime import datetime

# Timestamp pour nommage des fichiers de résultats
ts = datetime.now().strftime("%Y%m%d_%H%M%S")

# ── Config ─────────────────────────────────────────────────────────────────────
INPUT_DIR = os.path.join(_ROOT, "dev", "data", "input", "SL18in")
CSV_OUT   = os.path.join(_DIR, f"sweep2_results_sl18_{ts}.csv")
XLSX_OUT  = os.path.join(_DIR, f"sweep2_results_sl18_{ts}.xlsx")

# Espace de recherche Optuna
SEARCH_SPACE = {
    "multi_client_minimum_ratio": (0.02, 0.23),
    "mc_max_delta":               (0.01, 0.25),   # mc_max = mc_min + delta
    "min_filling_ratio":          (0.20, 0.55),
}

# ── Log parsing ────────────────────────────────────────────────────────────────

_RE_LNS_DONE = re.compile(
    r"\[LNS-(?P<kind>mono|multi)[^\]]*\] Done\.\s+"
    r"(?P<iters>\d+) iters in (?P<elapsed>[\d.]+)s"
    r" \| improvements: (?P<improvements>\d+)"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
    r" \| pallets: (?P<pal_before>\d+)→(?P<pal_after>\d+)"
)
_RE_PP_FILL_DONE = re.compile(
    r"\[Post[^\]]*\] fill equalization done — (?P<improvements>\d+) improvements"
    r" \((?P<skipped>\d+) skipped\) in (?P<elapsed>[\d.]+)s"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
)
_RE_PP_P2_DONE_NONE = re.compile(
    r"\[Post[^\]]*\] P2 done — no improvement found.*?"
    r"in (?P<elapsed>[\d.]+)s"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
)
_RE_PP_P2_DONE_OK = re.compile(
    r"\[Post[^\]]*\] P2 done — (?P<improvements>\d+) improvements"
    r" \((?P<skipped>\d+) skipped\) in (?P<elapsed>[\d.]+)s"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
)


def _parse_log(log: str) -> dict:
    stats: dict = {}

    mono_iters = mono_improvements = mono_stagnation = mono_elapsed = 0
    mono_pal_before = mono_pal_after = 0
    multi_iters = multi_improvements = multi_stagnation = multi_elapsed = 0
    multi_pal_before = multi_pal_after = 0

    for m in _RE_LNS_DONE.finditer(log):
        if m.group("kind") == "mono":
            mono_iters        += int(m.group("iters"))
            mono_improvements += int(m.group("improvements"))
            mono_stagnation   += int(m.group("stagnation"))
            mono_elapsed      += float(m.group("elapsed"))
            mono_pal_before   += int(m.group("pal_before"))
            mono_pal_after    += int(m.group("pal_after"))
        else:
            multi_iters        += int(m.group("iters"))
            multi_improvements += int(m.group("improvements"))
            multi_stagnation   += int(m.group("stagnation"))
            multi_elapsed      += float(m.group("elapsed"))
            multi_pal_before   += int(m.group("pal_before"))
            multi_pal_after    += int(m.group("pal_after"))

    stats["mono_iters"]        = mono_iters
    stats["mono_improvements"] = mono_improvements
    stats["mono_stagnation"]   = mono_stagnation
    stats["mono_stag_pct"]     = round(mono_stagnation / max(1, mono_iters) * 100, 1)
    stats["mono_elapsed_s"]    = round(mono_elapsed, 1)
    stats["mono_pal_delta"]    = mono_pal_before - mono_pal_after
    stats["multi_iters"]       = multi_iters
    stats["multi_improvements"]= multi_improvements
    stats["multi_stagnation"]  = multi_stagnation
    stats["multi_stag_pct"]    = round(multi_stagnation / max(1, multi_iters) * 100, 1)
    stats["multi_elapsed_s"]   = round(multi_elapsed, 1)
    stats["multi_pal_delta"]   = multi_pal_before - multi_pal_after

    fill_improvements = fill_stagnation = fill_elapsed = 0
    for m in _RE_PP_FILL_DONE.finditer(log):
        fill_improvements += int(m.group("improvements"))
        fill_stagnation   += int(m.group("stagnation"))
        fill_elapsed      += float(m.group("elapsed"))
    stats["pp_fill_improvements"] = fill_improvements
    stats["pp_fill_stagnation"]   = fill_stagnation
    stats["pp_fill_elapsed_s"]    = round(fill_elapsed, 1)

    p2_improvements = p2_stagnation = p2_elapsed = 0
    for m in _RE_PP_P2_DONE_OK.finditer(log):
        p2_improvements += int(m.group("improvements"))
        p2_stagnation   += int(m.group("stagnation"))
        p2_elapsed      += float(m.group("elapsed"))
    for m in _RE_PP_P2_DONE_NONE.finditer(log):
        p2_stagnation += int(m.group("stagnation"))
        p2_elapsed    += float(m.group("elapsed"))
    stats["pp_p2_improvements"] = p2_improvements
    stats["pp_p2_stagnation"]   = p2_stagnation
    stats["pp_p2_elapsed_s"]    = round(p2_elapsed, 1)

    return stats


# ── Pipeline ───────────────────────────────────────────────────────────────────

def _run_pipeline(params: OptimizationParameters) -> tuple:
    csv_files   = sorted(glob.glob(os.path.join(INPUT_DIR, "*.csv")))
    all_pallets = []
    combined    = []
    for csv_file in csv_files:
        buf        = io.StringIO()
        old_stdout = sys.stdout
        sys.stdout = buf
        try:
            boxes   = read_boxes_from_csv(csv_file)
            pallets = optimize_palletization(boxes, params)
            if params.enable_post_processing:
                pallets = postprocess(pallets, boxes, params)
            all_pallets.extend(pallets)
        finally:
            sys.stdout = old_stdout
        combined.append(buf.getvalue())
    return all_pallets, "".join(combined)


# ── Worker ─────────────────────────────────────────────────────────────────────

def _run_trial_worker(
    trial_number: int,
    mc_min: float,
    mc_max: float,
    fill:   float,
) -> dict:
    """Executed in a worker process. Returns a metrics dict."""
    baseline = OptimizationParameters()
    params   = copy.deepcopy(baseline)
    params.multi_client_minimum_ratio = mc_min
    params.multi_client_maximum_ratio = mc_max
    params.min_filling_ratio          = fill
    params.enable_post_processing     = False

    t0 = time.time()
    try:
        pallets, log = _run_pipeline(params)
        runtime = round(time.time() - t0, 1)
        stats   = _parse_log(log)
        active  = [p for p in pallets if len(p.boxes) > 0]
        n_pal   = len(active)
        n_multi = sum(1 for p in active if p.is_multi_client)
        n_mono  = n_pal - n_multi
        stats["final_pallets"]      = n_pal
        stats["n_multi"]            = n_multi
        stats["n_mono"]             = n_mono
        stats["multi_client_ratio"] = round(n_multi / max(1, n_pal), 4)
        error = None
    except Exception as exc:
        runtime = round(time.time() - t0, 1)
        stats   = {}
        error   = str(exc)

    return {
        "trial_number":               trial_number,
        "multi_client_minimum_ratio": mc_min,
        "multi_client_maximum_ratio": mc_max,
        "min_filling_ratio":          fill,
        "total_runtime_s":            runtime,
        "error":                      error,
        **stats,
    }


# ── Report writers ─────────────────────────────────────────────────────────────

_FIELDNAMES = [
    "trial_number", "pareto",
    "multi_client_minimum_ratio", "multi_client_maximum_ratio", "min_filling_ratio",
    "total_runtime_s", "final_pallets", "n_multi", "n_mono", "multi_client_ratio",
    "mono_iters", "mono_improvements", "mono_stagnation", "mono_stag_pct",
    "mono_elapsed_s", "mono_pal_delta",
    "multi_iters", "multi_improvements", "multi_stagnation", "multi_stag_pct",
    "multi_elapsed_s", "multi_pal_delta",
    "pp_fill_improvements", "pp_fill_stagnation", "pp_fill_elapsed_s",
    "pp_p2_improvements", "pp_p2_stagnation", "pp_p2_elapsed_s",
]

_SECTIONS = [
    ("Config",       10),   # trial + pareto + 3 params + runtime + total + n_multi + n_mono + ratio
    ("LNS Mono",      6),
    ("LNS Multi",     6),
    ("PP Fill",       3),
    ("PP P2",         3),
]


def _write_csv(rows: list) -> None:
    with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_FIELDNAMES, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"[Sweep2] CSV  → {CSV_OUT}")


def _write_xlsx(rows: list, pareto_numbers: set) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        print("[Sweep2] openpyxl non installé — XLSX ignoré.  pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pareto Results"

    hdr_font       = Font(bold=True, color="FFFFFF")
    sec_fill       = PatternFill("solid", fgColor="1F497D")
    col_fill       = PatternFill("solid", fgColor="2F4F8F")
    pareto_fill    = PatternFill("solid", fgColor="E2EFDA")   # vert clair
    pareto_font    = Font(bold=True)
    dominated_fill = PatternFill("solid", fgColor="F2F2F2")

    # Ligne 1 : groupes de colonnes
    col = 1
    for sec_title, sec_width in _SECTIONS:
        ws.merge_cells(
            start_row=1, start_column=col,
            end_row=1,   end_column=col + sec_width - 1,
        )
        cell           = ws.cell(row=1, column=col, value=sec_title)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = sec_fill
        cell.alignment = Alignment(horizontal="center")
        col += sec_width

    # Ligne 2 : noms de colonnes
    for c, name in enumerate(_FIELDNAMES, 1):
        cell           = ws.cell(row=2, column=c, value=name)
        cell.font      = hdr_font
        cell.fill      = col_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Données
    for r, row in enumerate(rows, 3):
        is_pareto = row.get("trial_number") in pareto_numbers
        fill = pareto_fill if is_pareto else dominated_fill
        font = pareto_font if is_pareto else Font()
        for c, key in enumerate(_FIELDNAMES, 1):
            val  = row.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.font = font

    # Largeurs automatiques
    for c, name in enumerate(_FIELDNAMES, 1):
        max_len = max(len(name), *(len(str(row.get(name, ""))) for row in rows))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 22)

    # Feuille Pareto uniquement
    ws2 = wb.create_sheet("Pareto Front")
    for c, name in enumerate(_FIELDNAMES, 1):
        cell           = ws2.cell(row=1, column=c, value=name)
        cell.font      = hdr_font
        cell.fill      = col_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    pareto_rows = sorted(
        [r for r in rows if r.get("trial_number") in pareto_numbers],
        key=lambda r: r.get("final_pallets", 9999),
    )
    for r, row in enumerate(pareto_rows, 2):
        for c, key in enumerate(_FIELDNAMES, 1):
            cell      = ws2.cell(row=r, column=c, value=row.get(key, ""))
            cell.fill = pareto_fill
            cell.font = pareto_font
    for c, name in enumerate(_FIELDNAMES, 1):
        max_len = max(len(name), *(len(str(r.get(name, ""))) for r in pareto_rows) if pareto_rows else [0])
        ws2.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 22)

    wb.save(XLSX_OUT)
    print(f"[Sweep2] XLSX → {XLSX_OUT}")


# ── Main ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Pareto sweep NSGA-II sur paramètres multi-client"
    )
    parser.add_argument("--trials",  type=int, default=80,
                        help="Nombre total d'essais Optuna (défaut: 80)")
    parser.add_argument("--workers", type=int, default=4,
                        help="Workers parallèles (défaut: 4)")
    parser.add_argument("--seed",    type=int, default=42,
                        help="Graine aléatoire NSGA-II (défaut: 42)")
    args = parser.parse_args()

    try:
        import optuna
        from optuna.samplers import NSGAIISampler
    except ImportError:
        print("[Sweep2] ERREUR : optuna non installé.  pip install optuna")
        sys.exit(1)

    optuna.logging.set_verbosity(optuna.logging.WARNING)

    study = optuna.create_study(
        directions=["minimize", "minimize"],
        sampler=NSGAIISampler(seed=args.seed),
        study_name="pallet_multi_pareto",
    )

    csv_count = len(sorted(glob.glob(os.path.join(INPUT_DIR, "*.csv"))))
    print(f"[Sweep2] Optuna NSGA-II  — {args.trials} trials, {args.workers} workers, seed={args.seed}")
    print(f"[Sweep2] Input dir : {INPUT_DIR}  ({csv_count} CSV)")
    print(f"[Sweep2] Objectifs : final_pallets (↓)   multi_client_ratio (↓)")
    print(f"[Sweep2] Espace    : mc_min [{SEARCH_SPACE['multi_client_minimum_ratio']}]  "
          f"mc_delta [{SEARCH_SPACE['mc_max_delta']}]  "
          f"fill [{SEARCH_SPACE['min_filling_ratio']}]\n")

    rows_by_number: dict = {}
    n_done = 0

    with ProcessPoolExecutor(max_workers=args.workers) as executor:
        submitted = 0
        while submitted < args.trials:
            batch = min(args.workers, args.trials - submitted)
            trials = [study.ask() for _ in range(batch)]

            futures = {}
            for trial in trials:
                mc_min   = trial.suggest_float(
                    "multi_client_minimum_ratio",
                    *SEARCH_SPACE["multi_client_minimum_ratio"],
                )
                mc_delta = trial.suggest_float(
                    "mc_max_delta",
                    *SEARCH_SPACE["mc_max_delta"],
                )
                fill     = trial.suggest_float(
                    "min_filling_ratio",
                    *SEARCH_SPACE["min_filling_ratio"],
                )
                # mc_max garanti > mc_min, plafonné à 0.50
                mc_max = min(round(mc_min + mc_delta, 4), 0.50)
                mc_min = round(mc_min, 4)
                fill   = round(fill,   4)

                futures[
                    executor.submit(_run_trial_worker, trial.number, mc_min, mc_max, fill)
                ] = trial

            for future in as_completed(futures):
                trial = futures[future]
                try:
                    row = future.result()
                except Exception as exc:
                    print(f"  ✗ trial {trial.number:3d}  EXCEPTION: {exc}")
                    study.tell(trial, (9999.0, 1.0))
                    continue

                obj1 = float(row.get("final_pallets",      9999))
                obj2 = float(row.get("multi_client_ratio", 1.0))
                study.tell(trial, (obj1, obj2))
                rows_by_number[trial.number] = row
                n_done += 1

                if row.get("error"):
                    print(f"  ✗ trial {trial.number:3d}  ERROR: {row['error']}")
                else:
                    print(
                        f"  ✓ trial {trial.number:3d}  "
                        f"pallets={obj1:5.0f}  "
                        f"ratio={obj2:.3f}  "
                        f"mc=[{row['multi_client_minimum_ratio']:.3f}, "
                        f"{row['multi_client_maximum_ratio']:.3f}]  "
                        f"fill={row['min_filling_ratio']:.3f}  "
                        f"t={row['total_runtime_s']}s  "
                        f"[{n_done}/{args.trials}]"
                    )

            submitted += batch

    # ── Résultats ──────────────────────────────────────────────────────────────
    pareto_numbers = {t.number for t in study.best_trials}

    all_rows = []
    for t in sorted(study.trials, key=lambda t: t.number):
        if t.number not in rows_by_number:
            continue
        row = rows_by_number[t.number]
        row["pareto"] = 1 if t.number in pareto_numbers else 0
        all_rows.append(row)

    # Résumé front de Pareto
    pareto_rows = sorted(
        [r for r in all_rows if r["pareto"]],
        key=lambda r: r.get("final_pallets", 9999),
    )
    print(f"\n{'='*70}")
    print(f"[Sweep2] Front de Pareto — {len(pareto_rows)} solutions sur {len(all_rows)} essais")
    print(f"  {'Trial':>6}  {'Pallets':>7}  {'Multi%':>7}  {'mc_min':>7}  {'mc_max':>7}  {'fill':>6}  {'runtime':>8}")
    print(f"  {'-'*6}  {'-'*7}  {'-'*7}  {'-'*7}  {'-'*7}  {'-'*6}  {'-'*8}")
    for r in pareto_rows:
        print(
            f"  {r['trial_number']:>6}  "
            f"{r.get('final_pallets', '?'):>7}  "
            f"{r.get('multi_client_ratio', 0)*100:>6.1f}%  "
            f"{r['multi_client_minimum_ratio']:>7.3f}  "
            f"{r['multi_client_maximum_ratio']:>7.3f}  "
            f"{r['min_filling_ratio']:>6.3f}  "
            f"{r['total_runtime_s']:>7.1f}s"
        )
    print(f"{'='*70}\n")

    _write_csv(all_rows)
    _write_xlsx(all_rows, pareto_numbers)
