"""
run_sweep.py — Parameter sweep for LNS time/iteration budgets.

Runs the full pipeline (phases 1-4 + post-processing) for each configuration
in SWEEP_GRID, captures stagnation stats from the console logs, and writes
sweep_results.csv + sweep_results.xlsx.

Usage:
    python run_sweep.py

Output:
    sweep_results.csv   — machine-readable results (written to experiments/ folder)
    sweep_results.xlsx  — formatted report for analysis (written to experiments/ folder)
"""

import copy
import glob
import io
import os
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Path setup ────────────────────────────────────────────────────────────────
_DIR  = os.path.dirname(os.path.abspath(__file__))            # .../pallet_optimizer/dev/experiments/
_ROOT = os.path.dirname(os.path.dirname(_DIR))                 # .../pallet_optimizer/ (git root)

import re
import csv

from pallet_optimizer.config.parameters import OptimizationParameters
from pallet_optimizer.file_io.csv_reader import read_boxes_from_csv
from pallet_optimizer.optimizer.pallet_optimizer import optimize_palletization
from pallet_optimizer.heuristics.post_processing import postprocess

# ── Input ─────────────────────────────────────────────────────────────────────
INPUT_DIR    = os.path.join(_ROOT, "dev", "data", "input", "SL18in")
CSV_OUT      = os.path.join(_DIR, "sweep_results.csv")
XLSX_OUT     = os.path.join(_DIR, "sweep_results.xlsx")
N_WORKERS    = 4   # configs tournées en parallèle

# ── Parameter grid ─────────────────────────────────────────────────────────────
# Each entry overrides only the specified keys; all others keep baseline values.
# Keys map directly to OptimizationParameters attribute names.
SWEEP_GRID = [
    # ── Référence ────────────────────────────────────────────────────────────────
    {"name": "baseline"},

    # ── Groupe A : mono — itérations par palette ──────────────────────────────────
    {"name": "mono_ip5",  "lns_mono_iter_per_pallet":  5},
    {"name": "mono_ip10", "lns_mono_iter_per_pallet": 10},
    {"name": "mono_ip20", "lns_mono_iter_per_pallet": 20},
    {"name": "mono_ip40", "lns_mono_iter_per_pallet": 40},

    # ── Groupe A2 : mono — temps par palette ──────────────────────────────────────
    {"name": "mono_tp03", "lns_mono_time_per_pallet": 0.3},
    {"name": "mono_tp07", "lns_mono_time_per_pallet": 0.7},   # défaut
    {"name": "mono_tp15", "lns_mono_time_per_pallet": 1.5},

    # ── Groupe B : multi — itérations par palette ─────────────────────────────────
    {"name": "multi_ip5",  "lns_multi_iter_per_pallet":  5},
    {"name": "multi_ip10", "lns_multi_iter_per_pallet": 10},
    {"name": "multi_ip20", "lns_multi_iter_per_pallet": 20},
    {"name": "multi_ip40", "lns_multi_iter_per_pallet": 40},

    # ── Groupe B2 : multi — temps par palette ─────────────────────────────────────
    {"name": "multi_tp02", "lns_multi_time_per_pallet": 0.2},
    {"name": "multi_tp05", "lns_multi_time_per_pallet": 0.5},  # défaut
    {"name": "multi_tp10", "lns_multi_time_per_pallet": 1.0},

    # ── Groupe C : post-processing — itérations par palette ───────────────────────
    {"name": "pp_ip5",  "pp_iter_per_pallet":  5},
    {"name": "pp_ip10", "pp_iter_per_pallet": 10},
    {"name": "pp_ip20", "pp_iter_per_pallet": 20},
    {"name": "pp_ip40", "pp_iter_per_pallet": 40},

    # ── Groupe C2 : post-processing — temps par palette ───────────────────────────
    {"name": "pp_tp02", "pp_time_per_pallet": 0.2},
    {"name": "pp_tp05", "pp_time_per_pallet": 0.5},  # défaut
    {"name": "pp_tp10", "pp_time_per_pallet": 1.0},

    # ── Config « slim » candidate — à ajuster après analyse ─────────────────────
    # {"name": "slim_candidate",
    #  "lns_mono_iter_per_pallet": 10, "lns_mono_time_per_pallet": 0.7,
    #  "lns_multi_iter_per_pallet": 10, "lns_multi_time_per_pallet": 0.5,
    #  "pp_iter_per_pallet": 20, "pp_time_per_pallet": 0.5},
]


# ── Log parsing ────────────────────────────────────────────────────────────────

_RE_LNS_DONE = re.compile(
    r"\[LNS-(?P<kind>mono|multi)[^\]]*\] Done\.\s+"
    r"(?P<iters>\d+) iters in (?P<elapsed>[\d.]+)s"
    r" \| improvements: (?P<improvements>\d+)"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
    r" \| pallets: (?P<pal_before>\d+)→(?P<pal_after>\d+)"
)

_RE_PP_FILL_DONE = re.compile(
    r"\[Post[^\]]*\] fill equalization done — (?P<improvements>\d+) improvements"
    r" \((?P<skipped>\d+) skipped\) in (?P<elapsed>[\d.]+)s"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
)
_RE_PP_P2_DONE_NONE = re.compile(
    r"\[Post[^\]]*\] P2 done — no improvement found.*?"
    r"in (?P<elapsed>[\d.]+)s"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
)
_RE_PP_P2_DONE_OK = re.compile(
    r"\[Post[^\]]*\] P2 done — (?P<improvements>\d+) improvements"
    r" \((?P<skipped>\d+) skipped\) in (?P<elapsed>[\d.]+)s"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
)

_RE_RESULT = re.compile(r"Result\s*:\s*(\d+) pallet\(s\)")


def _parse_log(log: str) -> dict:
    """Extracts metrics from captured stdout."""
    stats: dict = {}

    mono_iters = mono_improvements = mono_stagnation = mono_elapsed = 0
    mono_pal_before = mono_pal_after = 0
    for m in _RE_LNS_DONE.finditer(log):
        if m.group("kind") == "mono":
            mono_iters       += int(m.group("iters"))
            mono_improvements+= int(m.group("improvements"))
            mono_stagnation  += int(m.group("stagnation"))
            mono_elapsed     += float(m.group("elapsed"))
            mono_pal_before  += int(m.group("pal_before"))
            mono_pal_after   += int(m.group("pal_after"))

    stats["mono_iters"]        = mono_iters
    stats["mono_improvements"] = mono_improvements
    stats["mono_stagnation"]   = mono_stagnation
    stats["mono_stag_pct"]     = round(mono_stagnation / max(1, mono_iters) * 100, 1)
    stats["mono_elapsed_s"]    = round(mono_elapsed, 1)
    stats["mono_pal_delta"]    = mono_pal_before - mono_pal_after

    multi_iters = multi_improvements = multi_stagnation = multi_elapsed = 0
    multi_pal_before = multi_pal_after = 0
    for m in _RE_LNS_DONE.finditer(log):
        if m.group("kind") == "multi":
            multi_iters       += int(m.group("iters"))
            multi_improvements+= int(m.group("improvements"))
            multi_stagnation  += int(m.group("stagnation"))
            multi_elapsed     += float(m.group("elapsed"))
            multi_pal_before  += int(m.group("pal_before"))
            multi_pal_after   += int(m.group("pal_after"))

    stats["multi_iters"]        = multi_iters
    stats["multi_improvements"] = multi_improvements
    stats["multi_stagnation"]   = multi_stagnation
    stats["multi_stag_pct"]     = round(multi_stagnation / max(1, multi_iters) * 100, 1)
    stats["multi_elapsed_s"]    = round(multi_elapsed, 1)
    stats["multi_pal_delta"]    = multi_pal_before - multi_pal_after

    fill_improvements = fill_stagnation = fill_elapsed = 0
    for m in _RE_PP_FILL_DONE.finditer(log):
        fill_improvements += int(m.group("improvements"))
        fill_stagnation   += int(m.group("stagnation"))
        fill_elapsed      += float(m.group("elapsed"))

    stats["pp_fill_improvements"] = fill_improvements
    stats["pp_fill_stagnation"]   = fill_stagnation
    stats["pp_fill_elapsed_s"]    = round(fill_elapsed, 1)

    p2_improvements = p2_stagnation = p2_elapsed = 0
    for m in _RE_PP_P2_DONE_OK.finditer(log):
        p2_improvements += int(m.group("improvements"))
        p2_stagnation   += int(m.group("stagnation"))
        p2_elapsed      += float(m.group("elapsed"))
    for m in _RE_PP_P2_DONE_NONE.finditer(log):
        p2_stagnation += int(m.group("stagnation"))
        p2_elapsed    += float(m.group("elapsed"))

    stats["pp_p2_improvements"] = p2_improvements
    stats["pp_p2_stagnation"]   = p2_stagnation
    stats["pp_p2_elapsed_s"]    = round(p2_elapsed, 1)

    result_matches = _RE_RESULT.findall(log)
    stats["final_pallets"] = int(result_matches[-1]) if result_matches else 0

    return stats


# ── Pipeline ──────────────────────────────────────────────────────────────────

def _run_pipeline(params: OptimizationParameters) -> tuple[list, str]:
    """Runs the pipeline on every CSV in INPUT_DIR; returns (all_pallets, combined_stdout)."""
    csv_files   = sorted(glob.glob(os.path.join(INPUT_DIR, "*.csv")))
    all_pallets = []
    combined    = []

    for csv_file in csv_files:
        buf = io.StringIO()
        old_stdout = sys.stdout
        sys.stdout = buf
        try:
            boxes   = read_boxes_from_csv(csv_file)
            pallets = optimize_palletization(boxes, params)
            if params.enable_post_processing:
                pallets = postprocess(pallets, boxes, params)
            all_pallets.extend(pallets)
        finally:
            sys.stdout = old_stdout
        combined.append(buf.getvalue())

    return all_pallets, "".join(combined)


# ── Report writers ─────────────────────────────────────────────────────────────

_FIELDNAMES = [
    "config_name",
    "lns_mono_time_per_pallet", "lns_mono_iter_per_pallet",
    "lns_multi_time_per_pallet", "lns_multi_iter_per_pallet",
    "pp_time_per_pallet", "pp_iter_per_pallet",
    "total_runtime_s", "final_pallets",
    "mono_iters", "mono_improvements", "mono_stagnation", "mono_stag_pct",
    "mono_elapsed_s", "mono_pal_delta",
    "multi_iters", "multi_improvements", "multi_stagnation", "multi_stag_pct",
    "multi_elapsed_s", "multi_pal_delta",
    "pp_fill_improvements", "pp_fill_stagnation", "pp_fill_elapsed_s",
    "pp_p2_improvements", "pp_p2_stagnation", "pp_p2_elapsed_s",
]


def _write_csv(rows: list[dict]) -> None:
    with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_FIELDNAMES)
        w.writeheader()
        w.writerows(rows)
    print(f"[Sweep] CSV  → {CSV_OUT}")


def _write_xlsx(rows: list[dict]) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[Sweep] openpyxl not installed — skipping XLSX output.")
        print("        pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sweep Results"

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F4F8F")

    group_fills = {
        "baseline": PatternFill("solid", fgColor="D9E1F2"),
        "config":   PatternFill("solid", fgColor="F2F2F2"),
    }

    _SECTIONS = [
        ("Config", 7),
        ("Run", 2),
        ("LNS Mono", 6),
        ("LNS Multi", 6),
        ("PP Fill", 3),
        ("PP P2", 3),
    ]

    col = 1
    for sec_title, sec_width in _SECTIONS:
        ws.merge_cells(
            start_row=1, start_column=col,
            end_row=1,   end_column=col + sec_width - 1,
        )
        cell = ws.cell(row=1, column=col, value=sec_title)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor="1F497D")
        cell.alignment = Alignment(horizontal="center")
        col += sec_width

    for c, name in enumerate(_FIELDNAMES, 1):
        cell = ws.cell(row=2, column=c, value=name)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for r, row in enumerate(rows, 3):
        is_baseline = row["config_name"] == "baseline"
        fill = group_fills["baseline"] if is_baseline else group_fills["config"]
        for c, key in enumerate(_FIELDNAMES, 1):
            cell = ws.cell(row=r, column=c, value=row.get(key, ""))
            if is_baseline:
                cell.fill = fill
                cell.font = Font(bold=True)

    for c, name in enumerate(_FIELDNAMES, 1):
        max_len = max(len(name), *(len(str(row.get(name, ""))) for row in rows))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 25)

    ws.freeze_panes = "A3"
    wb.save(XLSX_OUT)
    print(f"[Sweep] XLSX → {XLSX_OUT}")


# ── Worker ────────────────────────────────────────────────────────────────────

def _run_config(cfg: dict) -> dict:
    """Execute one sweep config in a worker process; returns the result row + captured log."""
    baseline_params = OptimizationParameters()
    name   = cfg["name"]
    params = copy.deepcopy(baseline_params)
    for key, val in cfg.items():
        if key == "name":
            continue
        if hasattr(params, key):
            setattr(params, key, val)

    t0 = time.time()
    try:
        pallets, log = _run_pipeline(params)
        total_runtime = round(time.time() - t0, 1)
        stats = _parse_log(log)
        stats["final_pallets"] = sum(len(p.boxes) > 0 for p in pallets)
    except Exception as exc:
        total_runtime = round(time.time() - t0, 1)
        log = f"[Sweep] ERROR in {name}: {exc}\n"
        stats = {k: "ERROR" for k in _FIELDNAMES
                 if k not in ("config_name", "total_runtime_s")}

    return {
        "config_name":               name,
        "lns_mono_time_per_pallet":  params.lns_mono_time_per_pallet,
        "lns_mono_iter_per_pallet":  params.lns_mono_iter_per_pallet,
        "lns_multi_time_per_pallet": params.lns_multi_time_per_pallet,
        "lns_multi_iter_per_pallet": params.lns_multi_iter_per_pallet,
        "pp_time_per_pallet":        params.pp_time_per_pallet,
        "pp_iter_per_pallet":        params.pp_iter_per_pallet,
        "total_runtime_s":           total_runtime,
        "_log":                      log,
        **stats,
    }


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    _csv_files = sorted(glob.glob(os.path.join(INPUT_DIR, "*.csv")))
    print(f"[Sweep] Input dir : {INPUT_DIR}")
    print(f"[Sweep] CSV files : {len(_csv_files)}")
    print(f"[Sweep] Configs   : {len(SWEEP_GRID)}")
    print(f"[Sweep] Workers   : {N_WORKERS}\n")

    order = {cfg["name"]: i for i, cfg in enumerate(SWEEP_GRID)}
    rows  = [None] * len(SWEEP_GRID)

    with ProcessPoolExecutor(max_workers=N_WORKERS) as executor:
        futures = {executor.submit(_run_config, cfg): cfg["name"] for cfg in SWEEP_GRID}
        for future in as_completed(futures):
            name = futures[future]
            try:
                row = future.result()
            except Exception as exc:
                print(f"[Sweep] ✗ {name}: {exc}")
                continue

            log = row.pop("_log", "")
            print(f"\n{'='*60}")
            print(f"[Sweep] Config terminée : {name}")
            print(f"{'='*60}")
            print(log, end="")
            print(
                f"[Sweep] ✓ {name:20s}  runtime={row['total_runtime_s']}s  "
                f"pallets={row.get('final_pallets','?')}  "
                f"mono_stag={row.get('mono_stag_pct','?')}%  "
                f"multi_stag={row.get('multi_stag_pct','?')}%"
            )
            rows[order[name]] = row

    rows = [r for r in rows if r is not None]
    print(f"\n{'='*60}")
    print(f"[Sweep] {len(rows)}/{len(SWEEP_GRID)} configs terminées.")
    _write_csv(rows)
    _write_xlsx(rows)
