"""
kpi_writer.py — Génération des fichiers KPI depuis les objets Pallet en mémoire.

Produit sans aucune dépendance de visualisation (Dash, Plotly) :
    _kpi_cache.json     — métriques par palette, lues par le visualizer
    kpi_report_*.xlsx   — rapport Excel complet (openpyxl, optionnel)

Fonctions publiques :
    compute_kpi_rows(pallets)              → list[dict]
    load_kpi_cache(output_dir)             → dict
    save_kpi_cache(rows_by_file, dir)      → None
    write_excel(rows_by_file, output_dir)  → str (chemin ou "")
"""

import json
from datetime import datetime
from pathlib import Path
from typing import List

import numpy as np

from models.pallet import Pallet

_KPI_CACHE_FILE = "_kpi_cache.json"


class _NpEncoder(json.JSONEncoder):
    """Sérialise les types numpy (int64, float64, ndarray) en types Python natifs."""
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)


# ── Métriques géométriques ─────────────────────────────────────────────────────

def _surf_fill(boxes, pallet_length: float, pallet_width: float) -> float:
    """Taux de remplissage surfacique : fraction de l'empreinte palette couverte."""
    p_len = int(round(pallet_length))
    p_wid = int(round(pallet_width))
    p_area = p_len * p_wid
    if p_area <= 0:
        return 0.0
    grid = np.zeros((p_len, p_wid), dtype=bool)
    for b in boxes:
        x0 = max(0, int(b.x))
        x1 = min(p_len, int(round(b.x + b.length)))
        y0 = max(0, int(b.y))
        y1 = min(p_wid, int(round(b.y + b.width)))
        if x1 > x0 and y1 > y0:
            grid[x0:x1, y0:y1] = True
    return float(grid.sum()) / p_area


def _cog_offset(boxes, pallet_length: float, pallet_width: float) -> float:
    """Distance du centre de gravité au centre géométrique de la palette (cm)."""
    tw = sum(b.weight for b in boxes)
    if tw <= 0:
        return 0.0
    cx = sum(b.weight * (b.x + b.length / 2) for b in boxes) / tw
    cy = sum(b.weight * (b.y + b.width  / 2) for b in boxes) / tw
    dx = cx - pallet_length / 2
    dy = cy - pallet_width  / 2
    return (dx**2 + dy**2) ** 0.5


# ── Calcul des lignes KPI ──────────────────────────────────────────────────────

def compute_kpi_rows(pallets: List[Pallet]) -> list:
    """
    Calcule les métriques KPI par palette depuis les objets Pallet en mémoire.

    Retourne une liste de dicts (une entrée par palette non vide), triée par
    pallet.id. Chaque dict contient les clés attendues par le visualizer :
        pid, multi, clients, fill, surf_fill, cog, cog_x, cog_y, cog_z,
        height, p1, p2, weight, n_boxes, stability
    """
    rows = []
    for pallet in sorted(pallets, key=lambda p: p.id):
        boxes = pallet.boxes
        if not boxes:
            continue
        clients = sorted({b.client_id for b in boxes})
        tw = sum(b.weight for b in boxes)
        cog_x = sum(b.weight * (b.x + b.length / 2) for b in boxes) / tw if tw > 0 else 0.0
        cog_y = sum(b.weight * (b.y + b.width  / 2) for b in boxes) / tw if tw > 0 else 0.0
        cog_z = sum(b.weight * (b.z + b.height / 2) for b in boxes) / tw if tw > 0 else 0.0
        rows.append({
            "pid":       pallet.id,
            "multi":     pallet.is_multi_client,
            "clients":   clients,
            "fill":      pallet.volumetric_fill_ratio,
            "surf_fill": _surf_fill(boxes, pallet.length, pallet.width),
            "cog":       _cog_offset(boxes, pallet.length, pallet.width),
            "cog_x":     cog_x,
            "cog_y":     cog_y,
            "cog_z":     cog_z,
            "height":    pallet.current_height,
            "p1":        pallet.priority1_count,
            "p2":        pallet.priority2_count,
            "weight":    tw,
            "n_boxes":   len(boxes),
            "stability": pallet.worst_stability_ratio,
        })
    return rows


# ── Cache JSON ─────────────────────────────────────────────────────────────────

def load_kpi_cache(output_dir: Path) -> dict:
    """Charge _kpi_cache.json. Retourne {} si absent ou invalide."""
    path = output_dir / _KPI_CACHE_FILE
    try:
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def save_kpi_cache(rows_by_file: dict, output_dir: Path) -> None:
    """Écrit rows_by_file dans _kpi_cache.json (écrase l'existant)."""
    try:
        path = output_dir / _KPI_CACHE_FILE
        with open(path, "w", encoding="utf-8") as f:
            json.dump(rows_by_file, f, cls=_NpEncoder,
                      ensure_ascii=False, separators=(",", ":"))
    except Exception as e:
        print(f"[KPI] Avertissement : impossible d'écrire le cache KPI : {e}")


# ── Export Excel ───────────────────────────────────────────────────────────────

def write_excel(rows_by_file: dict, output_dir: Path) -> str:
    """
    Génère kpi_report_<ts>.xlsx dans output_dir.

    rows_by_file : dict { csv_filename → list[dict] } (format du cache KPI)
    Retourne le chemin du fichier créé, ou "" si openpyxl est absent ou si
    rows_by_file est vide.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[KPI] openpyxl non installé — Excel non généré.\n"
              "      Installez avec : pip install openpyxl")
        return ""

    if not rows_by_file:
        return ""

    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = str(output_dir / f"kpi_report_{ts}.xlsx")

    FMT_PCT   = "0.0%"
    FMT_FLOAT = "0.000"
    FMT_INT   = "0"

    wb         = openpyxl.Workbook()
    hdr_fill   = PatternFill("solid", fgColor="1E293B")
    sub_fill   = PatternFill("solid", fgColor="334155")
    title_fill = PatternFill("solid", fgColor="0F172A")
    hdr_font   = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=14)

    ws       = wb.active
    ws.title = "Rapport KPI"

    def _write_row(values, formats=None, *, header=False):
        ws.append(values)
        row_num = ws.max_row
        for col_idx, _ in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            if header:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(wrap_text=True, horizontal="center")
            elif formats and col_idx - 1 < len(formats) and formats[col_idx - 1]:
                cell.number_format = formats[col_idx - 1]

    # ── Totaux globaux ────────────────────────────────────────────────────────
    total_palettes = total_multi = total_boxes = total_p1 = total_p2 = 0
    fill_vals: list = []
    surf_vals: list = []
    for rows in rows_by_file.values():
        total_palettes += len(rows)
        total_multi    += sum(1 for r in rows if r["multi"])
        total_boxes    += sum(r["n_boxes"] for r in rows)
        total_p1       += sum(r["p1"] for r in rows)
        total_p2       += sum(r["p2"] for r in rows)
        fill_vals.extend(r["fill"]      for r in rows)
        surf_vals.extend(r["surf_fill"] for r in rows)

    avg_fill      = sum(fill_vals) / len(fill_vals) if fill_vals else 0.0
    avg_surf      = sum(surf_vals) / len(surf_vals) if surf_vals else 0.0
    ratio_glob    = total_p2 / total_p1 if total_p1 > 0 else 0.0
    multi_rate    = total_multi    / total_palettes if total_palettes else 0.0
    boxes_per_pal = total_boxes    / total_palettes if total_palettes else 0.0
    p1_per_pal    = total_p1       / total_palettes if total_palettes else 0.0
    p2_per_pal    = total_p2       / total_palettes if total_palettes else 0.0

    # ── En-tête + section globale ─────────────────────────────────────────────
    title_cell      = ws.cell(row=1, column=1, value="Rapport KPI")
    title_cell.fill = title_fill
    title_cell.font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    ws.append([])

    glob_cell      = ws.cell(row=ws.max_row + 1, column=1, value="KPIs Globaux")
    glob_cell.fill = sub_fill
    glob_cell.font = hdr_font
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=13)

    _write_row(["Fichiers analysés", "Total palettes", "Multi-client",
                "Taux multi", "Rempli. vol. moy.", "Rempli. surf. moy.",
                "Total Articles", "Articles / palette",
                "Total P1 (Meubles)", "P1 / palette",
                "Total P2 (Colis)", "P2 / palette", "Ratio P2/P1"], header=True)
    _write_row(
        [len(rows_by_file), total_palettes, total_multi,
         multi_rate, avg_fill, avg_surf,
         total_boxes, boxes_per_pal, total_p1, p1_per_pal, total_p2, p2_per_pal, ratio_glob],
        formats=([FMT_INT] * 3 + [FMT_PCT] * 3
                 + [FMT_INT, FMT_FLOAT] * 2 + [FMT_INT, FMT_FLOAT, FMT_FLOAT]),
    )
    ws.append([])
    ws.append([])

    # ── Section détail par fichier ────────────────────────────────────────────
    det_cell      = ws.cell(row=ws.max_row + 1, column=1, value="Détail par fichier")
    det_cell.fill = sub_fill
    det_cell.font = hdr_font
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=13)

    _write_row(["Fichier", "Palettes", "Multi-client", "Taux multi",
                "Rempli. vol.", "Rempli. surf.", "Moy. rempli. Mono", "Écart-type Mono",
                "Articles", "P1 (Meubles)", "P2 (Colis)", "Ratio P2/P1"], header=True)

    for fname, rows in rows_by_file.items():
        if not rows:
            continue
        n_total = len(rows)
        n_multi = sum(1 for r in rows if r["multi"])
        t_boxes = sum(r["n_boxes"] for r in rows)
        t_p1    = sum(r["p1"] for r in rows)
        t_p2    = sum(r["p2"] for r in rows)
        af      = sum(r["fill"]      for r in rows) / n_total
        asf     = sum(r["surf_fill"] for r in rows) / n_total
        rr      = t_p2 / t_p1 if t_p1 > 0 else 0.0
        mr      = n_multi / n_total if n_total else 0.0
        mono_fills = [r["fill"] for r in rows if not r["multi"]]
        mean_mono  = sum(mono_fills) / len(mono_fills) if mono_fills else None
        std_mono   = float(np.std(mono_fills))        if mono_fills else None
        _write_row(
            [fname, n_total, n_multi, mr, af, asf,
             mean_mono if mean_mono is not None else "",
             std_mono  if std_mono  is not None else "",
             t_boxes, t_p1, t_p2, rr],
            formats=[None, FMT_INT, FMT_INT, FMT_PCT, FMT_PCT, FMT_PCT,
                     FMT_PCT, FMT_PCT, FMT_INT, FMT_INT, FMT_INT, FMT_FLOAT],
        )

    ws.column_dimensions["A"].width = 42
    for idx in range(2, 14):
        ws.column_dimensions[get_column_letter(idx)].width = 16

    # ── Onglet 2 : détail par palette ─────────────────────────────────────────
    ws2 = wb.create_sheet(title="Détail par palette")
    pal_headers = [
        "Nom fichier", "Palette", "Client(s)",
        "Rempli. vol.", "Rempli. surf.", "Poids (kg)",
        "Colis", "P1", "P2", "Hauteur (cm)",
        "CdG X (cm)", "CdG Y (cm)", "CdG Z (cm)", "H / Rempli.", "Ratio stabilité",
    ]
    pal_formats = [
        None, None, None, FMT_PCT, FMT_PCT, FMT_FLOAT,
        FMT_INT, FMT_INT, FMT_INT,
        FMT_FLOAT, FMT_FLOAT, FMT_FLOAT, FMT_FLOAT, FMT_FLOAT, FMT_FLOAT,
    ]
    ws2.append(pal_headers)
    for col_idx in range(1, len(pal_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    def _fill_color(v):  return "16A34A" if v >= 0.6 else "F59E0B" if v >= 0.4 else "DC2626"
    def _surf_color(v):  return "0D9488" if v >= 0.6 else "F59E0B" if v >= 0.4 else "DC2626"
    def _stab_color(v):  return "DC2626" if v > 5.0 else "F59E0B" if v > 3.0 else "16A34A"

    for fname, rows in rows_by_file.items():
        if not rows:
            continue
        first_row = ws2.max_row + 1
        for r in rows:
            client_label = "Multi" if r["multi"] else str(r["clients"][0])
            h_over_fill  = (r["height"] / r["fill"]) if r["fill"] > 0 else None
            values = [
                fname, f"Palette {r['pid']}", client_label,
                r["fill"], r["surf_fill"], round(r["weight"], 1),
                r["n_boxes"], r["p1"], r["p2"], round(r["height"], 1),
                round(r["cog_x"], 1), round(r["cog_y"], 1), round(r["cog_z"], 1),
                round(h_over_fill, 0) if h_over_fill is not None else "",
                round(r["stability"], 3),
            ]
            ws2.append(values)
            row_num = ws2.max_row
            for col_idx, fmt in enumerate(pal_formats, start=1):
                if fmt:
                    ws2.cell(row=row_num, column=col_idx).number_format = fmt
            ws2.cell(row=row_num, column=4).font  = Font(color=_fill_color(r["fill"]))
            ws2.cell(row=row_num, column=5).font  = Font(color=_surf_color(r["surf_fill"]))
            if r["p2"] > 0:
                ws2.cell(row=row_num, column=9).font = Font(color="EA580C")
            ws2.cell(row=row_num, column=15).font = Font(color=_stab_color(r["stability"]))
        last_row = ws2.max_row
        if last_row > first_row:
            ws2.merge_cells(start_row=first_row, start_column=1,
                            end_row=last_row,    end_column=1)
            ws2.cell(row=first_row, column=1).alignment = Alignment(vertical="center")

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    for idx in range(4, len(pal_headers) + 1):
        ws2.column_dimensions[get_column_letter(idx)].width = 14
    ws2.freeze_panes = "B2"

    wb.save(excel_path)
    print(f"[KPI] Rapport Excel généré : {excel_path}")
    return excel_path
