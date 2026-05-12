"""
run_sweep.py — Parameter sweep for LNS time/iteration budgets.

Runs the full pipeline (phases 1-4 + post-processing) for each configuration
in SWEEP_GRID, captures stagnation stats from the console logs, and writes
sweep_results.csv + sweep_results.xlsx.

Usage:
    python run_sweep.py

Output:
    sweep_results.csv   — machine-readable results (written to test/ folder)
    sweep_results.xlsx  — formatted report for analysis (written to test/ folder)
"""

import copy
import io
import os
import sys
import time

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Path setup ────────────────────────────────────────────────────────────────
_DIR  = os.path.dirname(os.path.abspath(__file__))   # .../pallet_optimizer/test/
_BASE = os.path.dirname(_DIR)                         # .../pallet_optimizer/
if _BASE not in sys.path:
    sys.path.insert(0, _BASE)

import re
import csv

from config.parameters import OptimizationParameters
from file_io.csv_reader import read_boxes_from_csv
from optimizer.pallet_optimizer import optimize_palletization
from heuristics.post_processing import postprocess

# ── Input ─────────────────────────────────────────────────────────────────────
INPUT_CSV    = os.path.join(_BASE, r"input\tournee_type2026\tournee_type2026.csv")
OUTPUT_FAKE  = os.path.join(_BASE, r"output\sweep_run_results.csv")
CSV_OUT      = os.path.join(_DIR, "sweep_results.csv")
XLSX_OUT     = os.path.join(_DIR, "sweep_results.xlsx")

# ── Parameter grid ─────────────────────────────────────────────────────────────
# Each entry overrides only the specified keys; all others keep baseline values.
# Keys map directly to OptimizationParameters attribute names.
SWEEP_GRID = [
    # ── Référence ───────────────────────────────────────────────────────────────
    {"name": "baseline"},

    # ── Groupe A : mono LNS time budget ─────────────────────────────────────────
    {"name": "mono_t10",  "lns_mono_time_limit":  10.0},
    {"name": "mono_t20",  "lns_mono_time_limit":  20.0},
    {"name": "mono_t40",  "lns_mono_time_limit":  40.0},
    {"name": "mono_t60",  "lns_mono_time_limit":  60.0},
    {"name": "mono_t150", "lns_mono_time_limit": 150.0},

    # ── Groupe B : multi LNS time budget ────────────────────────────────────────
    {"name": "multi_t5",  "lns_multi_time_limit":  5.0},
    {"name": "multi_t10", "lns_multi_time_limit": 10.0},
    {"name": "multi_t20", "lns_multi_time_limit": 20.0},
    {"name": "multi_t60", "lns_multi_time_limit": 60.0},

    # ── Groupe C : post-processing budget ───────────────────────────────────────
    {"name": "pp_t5",     "pp_time_limit":  5.0, "pp_max_iterations": 125},
    {"name": "pp_t10",    "pp_time_limit": 10.0, "pp_max_iterations": 125},
    {"name": "pp_t40",    "pp_time_limit": 40.0, "pp_max_iterations": 500},

    # ── Groupe D : max_iterations (budget de convergence) ───────────────────────
    {"name": "mono_i100", "lns_mono_max_iterations": 100},
    {"name": "mono_i200", "lns_mono_max_iterations": 200},
    {"name": "multi_i100","lns_multi_max_iterations": 100},
    {"name": "multi_i150","lns_multi_max_iterations": 150},

    # ── Config « slim » — à renseigner après analyse des résultats ───────────────
    # {"name": "slim_candidate",
    #  "lns_mono_time_limit": 40.0, "lns_mono_max_iterations": 200,
    #  "lns_multi_time_limit": 20.0, "lns_multi_max_iterations": 150,
    #  "pp_time_limit": 10.0, "pp_max_iterations": 125},
]


# ── Log parsing ────────────────────────────────────────────────────────────────

# Patterns for the new "Done." lines:
#   [LNS-mono|client=X] Done. N iters in T.Ts | improvements: I | stagnation: S iters (P%) | pallets: A→B
#   [LNS-multi]         Done. N iters in T.Ts | improvements: I | stagnation: S iters (P%) | pallets: A→B
_RE_LNS_DONE = re.compile(
    r"\[LNS-(?P<kind>mono|multi)[^\]]*\] Done\.\s+"
    r"(?P<iters>\d+) iters in (?P<elapsed>[\d.]+)s"
    r" \| improvements: (?P<improvements>\d+)"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
    r" \| pallets: (?P<pal_before>\d+)→(?P<pal_after>\d+)"
)

# Pattern for PP phases:
#   [Post|…] fill equalization done — I improvements (S skipped) in T.Ts | stagnation: S iters (P%)
#   [Post|…] P2 done — I improvements (S skipped) in T.Ts | stagnation: S iters (P%) | Final cost=C
_RE_PP_FILL_DONE = re.compile(
    r"\[Post[^\]]*\] fill equalization done — (?P<improvements>\d+) improvements"
    r" \((?P<skipped>\d+) skipped\) in (?P<elapsed>[\d.]+)s"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
)
_RE_PP_P2_DONE = re.compile(
    r"\[Post[^\]]*\] P2 done — (?P<improvements>\d+|no improvement found, keeping original\.) ?"
    r"(?:improvements )?"
    r"\((?P<skipped>\d+) skipped\) in (?P<elapsed>[\d.]+)s"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
)
# Simpler P2 pattern for "no improvement" variant
_RE_PP_P2_DONE_NONE = re.compile(
    r"\[Post[^\]]*\] P2 done — no improvement found.*?"
    r"in (?P<elapsed>[\d.]+)s"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
)
_RE_PP_P2_DONE_OK = re.compile(
    r"\[Post[^\]]*\] P2 done — (?P<improvements>\d+) improvements"
    r" \((?P<skipped>\d+) skipped\) in (?P<elapsed>[\d.]+)s"
    r" \| stagnation: (?P<stagnation>\d+) iters \((?P<stag_pct>[\d.]+)%\)"
)

# Final pallets count from optimizer: "Result : N pallet(s)"
_RE_RESULT = re.compile(r"Result\s*:\s*(\d+) pallet\(s\)")


def _parse_log(log: str) -> dict:
    """Extracts metrics from captured stdout."""
    stats: dict = {}

    # LNS-mono passes (aggregate across all clients)
    mono_iters = mono_improvements = mono_stagnation = mono_elapsed = 0
    mono_pal_before = mono_pal_after = 0
    for m in _RE_LNS_DONE.finditer(log):
        if m.group("kind") == "mono":
            mono_iters       += int(m.group("iters"))
            mono_improvements+= int(m.group("improvements"))
            mono_stagnation  += int(m.group("stagnation"))
            mono_elapsed     += float(m.group("elapsed"))
            mono_pal_before  += int(m.group("pal_before"))
            mono_pal_after   += int(m.group("pal_after"))

    stats["mono_iters"]        = mono_iters
    stats["mono_improvements"] = mono_improvements
    stats["mono_stagnation"]   = mono_stagnation
    stats["mono_stag_pct"]     = round(mono_stagnation / max(1, mono_iters) * 100, 1)
    stats["mono_elapsed_s"]    = round(mono_elapsed, 1)
    stats["mono_pal_delta"]    = mono_pal_before - mono_pal_after

    # LNS-multi pass
    multi_iters = multi_improvements = multi_stagnation = multi_elapsed = 0
    multi_pal_before = multi_pal_after = 0
    for m in _RE_LNS_DONE.finditer(log):
        if m.group("kind") == "multi":
            multi_iters       += int(m.group("iters"))
            multi_improvements+= int(m.group("improvements"))
            multi_stagnation  += int(m.group("stagnation"))
            multi_elapsed     += float(m.group("elapsed"))
            multi_pal_before  += int(m.group("pal_before"))
            multi_pal_after   += int(m.group("pal_after"))

    stats["multi_iters"]        = multi_iters
    stats["multi_improvements"] = multi_improvements
    stats["multi_stagnation"]   = multi_stagnation
    stats["multi_stag_pct"]     = round(multi_stagnation / max(1, multi_iters) * 100, 1)
    stats["multi_elapsed_s"]    = round(multi_elapsed, 1)
    stats["multi_pal_delta"]    = multi_pal_before - multi_pal_after

    # PP — fill equalization (aggregate all groups)
    fill_improvements = fill_stagnation = fill_elapsed = 0
    fill_iters_total  = 0
    for m in _RE_PP_FILL_DONE.finditer(log):
        fill_improvements += int(m.group("improvements"))
        fill_stagnation   += int(m.group("stagnation"))
        fill_elapsed      += float(m.group("elapsed"))

    stats["pp_fill_improvements"] = fill_improvements
    stats["pp_fill_stagnation"]   = fill_stagnation
    stats["pp_fill_elapsed_s"]    = round(fill_elapsed, 1)

    # PP — P2 placement (aggregate all groups)
    p2_improvements = p2_stagnation = p2_elapsed = 0
    for m in _RE_PP_P2_DONE_OK.finditer(log):
        p2_improvements += int(m.group("improvements"))
        p2_stagnation   += int(m.group("stagnation"))
        p2_elapsed      += float(m.group("elapsed"))
    for m in _RE_PP_P2_DONE_NONE.finditer(log):
        p2_stagnation += int(m.group("stagnation"))
        p2_elapsed    += float(m.group("elapsed"))

    stats["pp_p2_improvements"] = p2_improvements
    stats["pp_p2_stagnation"]   = p2_stagnation
    stats["pp_p2_elapsed_s"]    = round(p2_elapsed, 1)

    # Final pallet count (last "Result : N pallet(s)" line)
    result_matches = _RE_RESULT.findall(log)
    stats["final_pallets"] = int(result_matches[-1]) if result_matches else 0

    return stats


# ── Pipeline ──────────────────────────────────────────────────────────────────

def _run_pipeline(params: OptimizationParameters) -> tuple[list, str]:
    """Runs the pipeline and returns (pallets, captured_stdout)."""
    buf = io.StringIO()
    old_stdout = sys.stdout
    sys.stdout = buf
    try:
        boxes   = read_boxes_from_csv(INPUT_CSV)
        pallets = optimize_palletization(boxes, params, output_path=OUTPUT_FAKE)
        if params.enable_post_processing:
            pallets = postprocess(pallets, boxes, params)
    finally:
        sys.stdout = old_stdout
    return pallets, buf.getvalue()


# ── Report writers ─────────────────────────────────────────────────────────────

_FIELDNAMES = [
    "config_name",
    "lns_mono_time", "lns_mono_iters_max",
    "lns_multi_time", "lns_multi_iters_max",
    "pp_time", "pp_iters_max",
    "total_runtime_s", "final_pallets",
    # Mono
    "mono_iters", "mono_improvements", "mono_stagnation", "mono_stag_pct",
    "mono_elapsed_s", "mono_pal_delta",
    # Multi
    "multi_iters", "multi_improvements", "multi_stagnation", "multi_stag_pct",
    "multi_elapsed_s", "multi_pal_delta",
    # PP fill
    "pp_fill_improvements", "pp_fill_stagnation", "pp_fill_elapsed_s",
    # PP P2
    "pp_p2_improvements", "pp_p2_stagnation", "pp_p2_elapsed_s",
]


def _write_csv(rows: list[dict]) -> None:
    with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_FIELDNAMES)
        w.writeheader()
        w.writerows(rows)
    print(f"[Sweep] CSV  → {CSV_OUT}")


def _write_xlsx(rows: list[dict]) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[Sweep] openpyxl not installed — skipping XLSX output.")
        print("        pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sweep Results"

    # Header style
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F4F8F")

    # Group fill colours
    group_fills = {
        "baseline": PatternFill("solid", fgColor="D9E1F2"),
        "config":   PatternFill("solid", fgColor="F2F2F2"),
    }

    # Section headers (multi-column spans)
    _SECTIONS = [
        ("Config", 7),
        ("Run", 2),
        ("LNS Mono", 6),
        ("LNS Multi", 6),
        ("PP Fill", 3),
        ("PP P2", 3),
    ]

    col = 1
    for sec_title, sec_width in _SECTIONS:
        ws.merge_cells(
            start_row=1, start_column=col,
            end_row=1,   end_column=col + sec_width - 1,
        )
        cell = ws.cell(row=1, column=col, value=sec_title)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor="1F497D")
        cell.alignment = Alignment(horizontal="center")
        col += sec_width

    # Column headers (row 2)
    for c, name in enumerate(_FIELDNAMES, 1):
        cell = ws.cell(row=2, column=c, value=name)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for r, row in enumerate(rows, 3):
        is_baseline = row["config_name"] == "baseline"
        fill = group_fills["baseline"] if is_baseline else group_fills["config"]
        for c, key in enumerate(_FIELDNAMES, 1):
            cell = ws.cell(row=r, column=c, value=row.get(key, ""))
            if is_baseline:
                cell.fill = fill
                cell.font = Font(bold=True)

    # Auto-width columns
    for c, name in enumerate(_FIELDNAMES, 1):
        max_len = max(len(name), *(len(str(row.get(name, ""))) for row in rows))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 25)

    ws.freeze_panes = "A3"

    wb.save(XLSX_OUT)
    print(f"[Sweep] XLSX → {XLSX_OUT}")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"[Sweep] Input : {INPUT_CSV}")
    print(f"[Sweep] Configs: {len(SWEEP_GRID)}\n")

    baseline_params = OptimizationParameters()
    rows = []

    for cfg in SWEEP_GRID:
        name = cfg["name"]
        print(f"\n{'='*60}")
        print(f"[Sweep] Config: {name}")
        print(f"{'='*60}")

        # Build params for this config
        params = copy.deepcopy(baseline_params)
        for key, val in cfg.items():
            if key == "name":
                continue
            if hasattr(params, key):
                setattr(params, key, val)
            else:
                print(f"[Sweep] WARNING: unknown parameter '{key}' — skipped.")

        t0 = time.time()
        try:
            pallets, log = _run_pipeline(params)
            total_runtime = round(time.time() - t0, 1)

            # Print captured log so progress is visible
            print(log, end="")

            stats = _parse_log(log)
            # Override final_pallets with direct count (more reliable)
            stats["final_pallets"] = sum(len(p.boxes) > 0 for p in pallets)

        except Exception as exc:
            total_runtime = round(time.time() - t0, 1)
            print(f"[Sweep] ERROR: {exc}")
            stats = {k: "ERROR" for k in _FIELDNAMES
                     if k not in ("config_name", "total_runtime_s")}

        row = {
            "config_name":        name,
            "lns_mono_time":      params.lns_mono_time_limit,
            "lns_mono_iters_max": params.lns_mono_max_iterations,
            "lns_multi_time":     params.lns_multi_time_limit,
            "lns_multi_iters_max":params.lns_multi_max_iterations,
            "pp_time":            params.pp_time_limit,
            "pp_iters_max":       params.pp_max_iterations,
            "total_runtime_s":    total_runtime,
            **stats,
        }
        rows.append(row)

        print(
            f"[Sweep] ✓ {name:20s}  runtime={total_runtime}s  "
            f"pallets={stats.get('final_pallets','?')}  "
            f"mono_stag={stats.get('mono_stag_pct','?')}%  "
            f"multi_stag={stats.get('multi_stag_pct','?')}%"
        )

    print(f"\n{'='*60}")
    print(f"[Sweep] All {len(rows)} configs completed.")
    _write_csv(rows)
    _write_xlsx(rows)
